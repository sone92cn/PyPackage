import xlrd, openpyxl
import os, re, threading
from copy import copy
from tkinter import filedialog
from tempfile import gettempdir
from datetime import date, datetime
from FelixFunc import getPosList, wipeNone, cleanedFileName, cleanedSheetName, createStringFromVarList
from openpyxl.styles import PatternFill, Border, Alignment, Font

base_date = date(1899, 12, 30)
base_time = datetime(1899, 12, 30)
type_date = type(base_date)
type_time = type(base_time)

class Cell(object):
    def __init__(self, file_type, sheet, row, col, merged=False, none=False):
        if none:
            self.value = None
            self.type = 'n'
            self.is_date = False
            self.merged = False
            self.font = Font()
            self.fill = PatternFill()
            self.border = Border()
            self.alignment = Alignment()
            self.number_format = 'General'
        elif file_type:
            cel = sheet.cell(row=row+1, column=col+1)
            value = cel.value
            if type(value) == type_date:
                dela = value - base_date
                value = dela.days
                self.is_date = True
            elif type(value) == type_time:
                dela = value - base_time
                value = dela.days
                self.is_date = True
            elif cel.data_type == 'n':
                if value != None:
                    value = round(value, 4)
                self.is_date = cel.is_date
            else:
                self.is_date = cel.is_date
            self.value = value
            self.type = cel.data_type
            self.merged = merged
            self.font = cel.font
            self.fill = cel.fill
            self.border = cel.border
            self.alignment = cel.alignment
            self.number_format = cel.number_format
            self.row = row
            self.column = col
        else:
            inx = sheet.cell_type(row, col)
            if inx == 3:
                value = int(sheet.cell_value(row, col))
                self.is_date = True
            elif inx == 2:
                value = round(value, 4)
                self.is_date = False
            else:
                value = Static.replaceNullStringWithNone(sheet.cell_value(row, col))
                self.is_date = False
            self.value = value
            self.type = Static.getCellTypeFromIndex(inx)
            self.merged = merged
            self.font = Font()
            self.fill = PatternFill()
            self.border = Border()
            self.alignment = Alignment()
            self.number_format = 'General'
            self.row = row
            self.column = col
    def __repr__(self):
        return "Cell:value=" + str(self.value) + ',type=' + str(self.type) + ',is_date=' + str(self.is_date)
    def set(self, cvalue=None, ctype='n', is_date=False):
        self.value = cvalue
        self.type = ctype
        self.is_date = is_date
        return True
    def write(self, cel):
        cel.data_type = self.type
        if self.type == "f":
            if "!" in self.value:
                cel.value = "Function Error"
                cel.data_type = "s"
            else:
                value = self.value
                for add in Cell.addre.findall(value):
                    off = Static.getCellOffset(self, add)
                    value = value.replace(add, cel.offset(off[0], off[1]).coordinate) 
                cel.value = value
        else:
            cel.value = self.value
        cel.font = copy(self.font)
        cel.fill = copy(self.fill)
        cel.border = copy(self.border)
        cel.alignment = copy(self.alignment)
        cel.number_format = self.number_format
        return True
    @classmethod
    def classInit(cls):
        cls.addre = re.compile(r'\$?[A-Z]+\$?[1-9]+', re.IGNORECASE)
    
class Static():
    @staticmethod
    def isCordnateAddress(cord):
        cord = cord.upper()
        for i, c in enumerate(cord):
            if ord(c) < ord('A') or ord(c) > ord('Z'):
                break
        else:
            return False
        
        if 0 < i < len(cord):
            for c in cord[i:]:
                if ord(c) < ord('0') or ord(c) > ord('9'):
                    return False
            return True
        else:
            return False
    @staticmethod
    def getCellTypeFromIndex(itype):
        if itype == 1:
            return 's'
        elif itype == 4:
            return 'b'
        elif itype == 5:
            return 'e'
        elif itype == 0 or itype == 2 or itype == 3 or itype == 6:
            return 'n'
    @staticmethod
    def getCellCordinateFromIndex(row, col):
        tmp = 0
        cord = ''
        row , col = row + 1, col + 1
        while col > 26:
            tmp = col % 26
            if tmp:
                cord = chr(tmp + 64) + cord
                col = col // 26
            else:
                cord = 'Z' + cord
                col = (col // 26) - 1
        if col:
            cord = chr(col + 64) + cord
        return cord + str(row)
    @staticmethod
    def getCellIndexFromCordinate(cord):
        cord = cord.replace("$", "")
        cord = cord.upper()
        for i, c in enumerate(cord):
            if ord(c) < 65:
                break
        else:
            raise Exception('Invalid Address String!')
        
        col, num = 0, 1
        cstr = cord[0:i]
        for c in cstr[::-1]:
            col =  col + (ord(c) - 64) * num
            num = num * 26
        row = int(cord[i:])
        return (row-1, col-1)
    @staticmethod
    def replaceNullStringWithNone(value):
        if type(value) == type('s'):
            if value == '':
                return None
        return value
    @staticmethod
    def getCellOffset(cel, addr):
        inx = Static.getCellIndexFromCordinate(addr)
        return (inx[0]-cel.row, inx[1]-cel.column)
            
class ExcelReader():
    groups = None
    def __init__(self, file_name, group='reader', data_only=True):
        if not os.path.isfile(file_name):
            raise Exception('File Not Existed!')
        ext = os.path.splitext(file_name)[1].lower()
        if ext == '.xls':
            self.type = 0
            self.file = file_name
            self.book = xlrd.open_workbook(file_name, formatting_info=True)
            self.sheets = self.book.sheets()
            self.sheet_names = self.book.sheet_names()
            self.valid_sheets = ExcelReader.loadVailidSheets(self, group)
            self.valid_sheet_names = [sheet.name for sheet in self.valid_sheets]
        elif ext == '.xlsx':
            self.type = 1
            self.file = file_name
            self.book = openpyxl.load_workbook(file_name, read_only=False, data_only=data_only, guess_types=True)
            self.sheets = self.book.worksheets
            self.sheet_names = self.book.sheetnames
            self.valid_sheets = ExcelReader.loadVailidSheets(self, group)
            self.valid_sheet_names = [self.sheet_names[self.book.index(sheet)] for sheet in self.valid_sheets]
        elif ext == '.xlsm':
            self.type = 1
            self.file = file_name
            self.book = openpyxl.load_workbook(file_name, read_only=False, data_only=data_only, guess_types=True, keep_vba=False)
            self.sheets = self.book.worksheets
            self.sheet_names = self.book.sheetnames
            self.valid_sheets = ExcelReader.loadVailidSheets(self, group)
            self.valid_sheet_names = [self.sheet_names[self.book.index(sheet)] for sheet in self.valid_sheets]
        else:
            raise Exception('Unknown File!')
    def __eq__(self, other):
        return self is other
    def __isMergedCell(self, sheet, row, col, merged_range=None):  #set top cell address to top_address if not None
        if self.type:
            if merged_range == None:
                if Static.getCellCordinateFromIndex(row, col) in sheet.merged_cells:
                    return True
            else:
                for crange in sheet.merged_cell_ranges:
                    cords = crange.split(':')
                    rlo, clo = Static.getCellIndexFromCordinate(cords[0])
                    rhi, chi = Static.getCellIndexFromCordinate(cords[1])
                    if rlo <= row <= rhi and clo <= col <= chi:
                        merged_range.append(rlo)
                        merged_range.append(rhi+1)
                        merged_range.append(clo)
                        merged_range.append(chi+1)
                        return True
        else:
            for crange in sheet.merged_cells:
                rlo, rhi, clo, chi = crange
                if rlo <= row < rhi and clo <= col < chi:
                    if merged_range != None:
                        merged_range.extend(crange)
                    return True
        return False
    def getMergedRanges(self, sheet, r_min=0, c_min=0, r_max=0, c_max=0, relative=False):
        ranges = []
        if self.type:
            for crange in sheet.merged_cell_ranges:
                cords = crange.split(':')
                rlo, clo = Static.getCellIndexFromCordinate(cords[0])
                rhi, chi = Static.getCellIndexFromCordinate(cords[1])
                rhi = rhi + 1
                chi = chi + 1
                if r_min and rlo < r_min:
                    rlo = r_min
                if c_min and clo < c_min:
                    clo = c_min
                if r_max and rhi > r_max:
                    rhi = r_max
                if c_max and chi > c_max:
                    chi = c_max
                if chi > clo and rhi > rlo:
                    if chi - clo + rhi - rlo > 2:
                        if relative:
                            ranges.append((rlo-r_min, clo-c_min, rhi-r_min, chi-c_min))
                        else:
                            ranges.append((rlo, clo, rhi, chi))
        else:
            for crange in sheet.merged_cells:
                rlo, rhi, clo, chi = crange
                if r_min and rlo < r_min:
                    rlo = r_min
                if c_min and clo < c_min:
                    clo = c_min
                if r_max and rhi > r_max:
                    rhi = r_max
                if c_max and chi > c_max:
                    chi = c_max
                if chi > clo and rhi > rlo:
                    if chi - clo + rhi - rlo > 2:
                        if relative:
                            ranges.append((rlo-r_min, clo-c_min, rhi-r_min, chi-c_min))
                        else:
                            ranges.append((rlo, clo, rhi, chi))
        return ranges
    def getBookPath(self):
        if self.file != '':
            return os.path.split(self.file)[0]
        else:
            return ''
    def getBookName(self, cut_ext=False):
        if self.file != '':
            book = os.path.split(self.file)[1]
            if cut_ext:
                return os.path.splitext(book)[0]
            else:
                return book
        else:
            return ''
    def getFullName(self):
        return self.file
    def getSheetName(self, sheet):
        if self.type:
            return sheet.title
        else:
            return sheet.name
    def getSheetObject(self, sheet_name):
        return self.sheets[self.sheet_names.index(sheet_name)]
    def getTotalCols(self, sheet):
        if self.type:
            return sheet.max_column
        else:
            return sheet.ncols
    def getTotalRows(self, sheet):
        if self.type:
            return sheet.max_row
        else:
            return sheet.nrows
    def getCellValue(self, sheet, row, col, check_merge=True, return_top=True, collect_all=False):
        if check_merge:
            rng = []
            merged = self.__isMergedCell(sheet, row, col, rng)
            if merged and return_top:
                row, col = rng[0], rng[2]
        else:
            merged = False
        if collect_all:
            return Cell(self.type, sheet, row, col, merged)
        else:
            if self.type:
                value = sheet.cell(row=row+1, column=col+1).value
                ctype = sheet.cell(row=row+1, column=col+1).data_type
                if type(value) == type_date:
                    dela = value - base_date
                    value = dela.days
                elif type(value) == type_time:
                    dela = value - base_time
                    value = dela.days
                elif ctype == 'n':
                    if value != None:
                        value = round(value, 5)
            else:
                value = sheet.cell_value(row, col)
                ctype = sheet.cell_type(row, col)
                if ctype == 3:
                    value = int(value)
                elif ctype == 2:
                    value = round(value, 5)
                else:
                    value = Static.replaceNullStringWithNone(value)
            return value
    def getColValues(self, sheet, col, start=0, check_merge=True, return_top=True, collect_all=False):
        return [self.getCellValue(sheet, i, col, check_merge, return_top, collect_all) for i in range(start, self.getTotalRows(sheet))]
    def getRowValues(self, sheet, row, start=0, check_merge=True, return_top=True, collect_all=False):
        return [self.getCellValue(sheet, row, i, check_merge, return_top, collect_all) for i in range(start, self.getTotalCols(sheet))]
    def getRowValuesByPos(self, sheet, row, pos, check_merge=True, return_top=True, collect_all=False, allow_error=False):
        if allow_error:
            values = []
            for i in pos:
                if i > -1:
                    values.append(self.getCellValue(sheet, row, i, check_merge, return_top, collect_all))
                else:
                    if collect_all:
                        values.append(Cell(None, None, None, None, none=True))
                    else:
                        values.append(None)
            return values
        else:
            return [self.getCellValue(sheet, row, i, check_merge, return_top, collect_all) for i in pos]
    def getRowsValuesByPos(self, sheet, pos, start_row=0, check_merge=True, return_top=True, collect_all=False, skip_none=False, skip_allnone=True, skip_sumrow=True):
        result = {}
        for i in range(start_row, self.getTotalRows(sheet)):
            infos = self.getRowValuesByPos(sheet, i, pos, check_merge, return_top, collect_all)
            if collect_all:
                values = [info.value for info in infos]
            else:
                values = infos
            if skip_none:
                if None in values:
                    continue
            if skip_allnone:
                if [None] * len(values) == values:
                    continue
            if skip_sumrow:
                for value in values:
                    if type(value) == type('s'):
                        if ExcelReader.sumre.match(value) != None:
                            break
                else:
                    result[i] = infos
            else:
                result[i] = infos
        return result
    def getRowPosByValues(self, sheet, row, check_list, check_merge=False, return_top=False, collect_all=False, all_included=True):
        values = self.getRowValues(sheet, row, start=0, check_merge=check_merge, return_top=return_top, collect_all=return_top)
        pos_list = []
        if getPosList(values, check_list, pos_list, all_included):
            return pos_list
        else:
            return []
    def getTitleRow(self, sheet, check_list, check_pos=[], guess_row=-1, max_row=9): #check_pos must be a empty list
        if guess_row > -1:
            values = self.getRowValues(sheet, guess_row, start=0, check_merge=False, return_top=False, collect_all=False)
            pos_list = []
            if getPosList(values, check_list, pos_list, True):
                check_pos.extend(pos_list)
                return guess_row
        for i in range(0, max_row):
            if i == guess_row:
                continue
            values = self.getRowValues(sheet, i, start=0, check_merge=False, return_top=False, collect_all=False)
            pos_list = []
            if getPosList(values, check_list, pos_list, True):
                check_pos.extend(pos_list)
                return guess_row
        raise Exception('Title row not found!')
    def guessTitleRow(self, sheet, allow_jump=False, not_merge=True, must_unique=True, max_row=9):
        for i in range(0, 9):
            infos = self.getRowValues(sheet, i, start=0, check_merge=True, return_top=False, collect_all=True)
            if not_merge:
                if set(info.merged for info in infos) != set([False]):
                    continue
            if allow_jump:
                for info in infos:
                    if info.value == None:
                        info.type = 's'
            if set(info.type for info in infos) != set(['s']):
                continue
            values = [info.value for info in infos]
            values = wipeNone(values)
            if values == []:
                continue
            if must_unique:
                if len(values) == len(set(values)):
                    return i
                else:
                    continue
            else:
                return i
        raise Exception('Guess error')
    @classmethod
    def classInit(cls):
        cls.sumre = re.compile(r'^.*((小|合|共|总) *计)|(汇 *总) *$', re.IGNORECASE)
        cls.groups = {'lock':threading.Lock(), 'invalid':[]}
    @classmethod
    def resetGroup(cls, group):
        cls.groups['lock'].acquire()
        cls.groups[group] = []
        cls.groups['lock'].release()
    @classmethod
    def getGroupShet(cls, group, sheet):
        if group in cls.groups:
            for shet in cls.groups[group]:
                if shet['sheet'] is sheet:
                    return shet
        return None
    @classmethod
    def getExcelObjects(cls, shets):
        objs = []
        if len(shets):
            objs = [shets[0]['xls']]
            for shet in shets[1:]:
                for xls in objs:
                    if xls is shet['xls']:
                        break
                else:
                    objs.append(shet['xls'])
        return objs
    @classmethod
    def setGroupSkip(cls, group, skip=True):
        cls.groups[group]['lock'].acquire()
        if group in cls.groups:
            for shet in cls.groups[group]:
                shet['search_skip']=skip
        cls.groups[group]['lock'].release()
    @classmethod
    def loadValidSheet(cls, xls, sheet, group, must_add=True):
        cls.groups['lock'].acquire()
        if not (must_add):
            if xls.getTotalCols(sheet) < 2 and xls.getTotalRows(sheet) < 2:
                return False
        if not (group in cls.groups):
            cls.groups[group] = [{'sheet':sheet, 'xls':xls, 'lock':threading.Lock(), 'valid':True, 'info':xls.getSheetName(sheet)+'<'+xls.getBookName()+'>'}]
        else:
            cls.groups[group].append({'sheet':sheet, 'xls':xls, 'lock':threading.Lock(), 'valid':True, 'info':xls.getSheetName(sheet)+'<'+xls.getBookName()+'>'})
        cls.groups['lock'].release()
        return True
    @classmethod
    def loadVailidSheets(cls, xls, group, must_add=False):
        no_added = True
        cur_group = []
        for sheet in xls.sheets:
            if xls.getTotalCols(sheet)>1 or xls.getTotalRows(sheet)>1:
                cur_group.append({'sheet':sheet, 'xls':xls, 'lock':threading.Lock(), 'valid':True, 'info':xls.getSheetName(sheet)+'<'+xls.getBookName()+'>'})
                no_added = False
        if must_add and no_added:
            cur_group.append({'sheet':xls.sheets[0], 'xls':xls, 'lock':threading.Lock(), 'valid':True, 'info':xls.getSheetName(xls.sheets[0])+'<'+xls.getBookName()+'>'})
        if len(cur_group):
            cls.groups['lock'].acquire()
            if not (group in cls.groups):
                cls.groups[group] = []
            cls.groups[group].extend(cur_group)
            cls.groups['lock'].release()
            return [item['sheet'] for item in cur_group]
        else:
            return []
    @staticmethod
    def getPosByRelativeRanges(cranges, r_min=0, c_min=0):
        poslist = []
        for crange in cranges:
            rlo, clo, rhi, chi = crange
            poslist.append((rlo+r_min, clo+c_min, rhi+r_min, chi+c_min))
        return poslist

class ExcelWriter(ExcelReader):
    writers = None
    def __init__(self, file_name='', group='writer', data_only=True):
        if file_name == '':
            self.book = openpyxl.Workbook()
        else:
            if not os.path.isfile(file_name):
                raise Exception('File Not Existed!')
            ext = os.path.splitext(file_name)[1].lower()
            if ext == '.xlsx':
                self.book = openpyxl.load_workbook(file_name, read_only=False, data_only=data_only, guess_types=True)
            elif ext == '.xlsm':
                self.book = openpyxl.load_workbook(file_name, read_only=False, data_only=data_only, guess_types=True, keep_vba=True)
            else:
                raise Exception('Unknown File!')
        self.type = 1
        self.file = file_name
        self.sheets = self.book.worksheets
        self.sheet_names = self.book.sheetnames
        self.valid_sheets = ExcelWriter.loadVailidSheets(self, group, True)
        self.valid_sheet_names = [sheet.title for sheet in self.valid_sheets]
    def setFileName(self, file_name):
        dpath = os.path.join(gettempdir(), 'python_excel_tool')
        if not os.path.isdir(dpath):
            os.mkdir(dpath)
        fpath = os.path.join(dpath, file_name)
        try:
            self.saveas(fpath)
        except:
            raise Exception('Save Error!')
        else:
            self.file = file_name
            try:
                os.remove(fpath)
            except:
                print('Fail to remove:' + fpath)
    def createSheet(self, title=None, index=None):
        return self.book.create_sheet(title, index)
    def mergeRanges(self, sheet, cranges):
        for crange in cranges:
            rlo, clo, rhi, chi = crange
            sheet.merge_cells(start_row=rlo+1, start_column=clo+1, end_row=rhi, end_column=chi)
            #sheet.merge_cells(Static.getCellCordinateFromIndex(rlo, clo) + ":" + Static.getCellCordinateFromIndex(rhi-1, chi-1))
    def updateCellValue(self, sheet, row, col, value, update_all=False):
        if update_all:
            return value.write(sheet.cell(row=row+1, column=col+1))
        else:
            sheet.cell(row=row+1, column=col+1).value = value
            return True
    def updateRowValues(self, sheet, row, values, start=0, update_all=False):
        for i, value in zip(range(start, start+len(values)), values):
            self.updateCellValue(sheet, row, i, value, update_all)
    def updateRowValuesByPos(self, sheet, row, pos, values, update_all=False):
        for i, value in zip(pos, values):
            self.updateCellValue(sheet, row, i, value, update_all)
    def appendRowValues(self, sheet, values, start=0, update_all=False):
        row = self.getTotalRows(sheet)
        for i, value in zip(range(start, start+len(values)), values):
            self.updateCellValue(sheet, row, i, value, update_all)
    def save(self):
        if self.file == '':
            file_name = filedialog.asksaveasfilename(title="另存为", filetypes=(("Excel Files", "*.xlsx"), ))
            if len(file_name):
                if os.path.splitext(file_name)[1].lower() != '.xlsx':
                    file_name += '.xlsx'
                self.file = file_name
                self.book.save(self.file)
            else:
                raise Exception('No file name to save.')
        else:
            self.book.save(self.file)
    def saveas(self, file_name):
        ext = os.path.splitext(file_name)[1].lower()
        if ext in ('.xlsx', '.xlsm'):
            self.book.save(file_name)
            self.file = file_name
        else:
            raise Exception('Invalid file name!')
    @classmethod
    def classInit(cls):
        cls.newid = 0
        cls.book = None
        cls.sumre = ExcelReader.sumre
        cls.groups = {'lock':threading.Lock(), 'invalid':[]}
        cls.writers = {'lock':threading.Lock()}
        cls.shtre = re.compile(r'^sheet[0-9]*$', re.IGNORECASE)
    @classmethod
    def getWriter(cls, writer, new_mode=None, new_model=None, not_copy=[], **auto_set):
        cls.writers['lock'].acquire()
        if not(writer in cls.writers):
            if new_mode == None:
                cls.writers['lock'].release()
                return None
            elif new_mode == 0:
                if cls.book == None:
                    cls.book = ExcelWriter(group='default_book')
                    sheet = cls.book.sheets[0]
                    cls.writers[writer] = [ExcelWriter.getGroupShet('default_book', sheet)]
                else:
                    sheet = cls.book.book.create_sheet()
                    cls.loadValidSheet(cls.book, sheet, group='default_book', must_add=True)
                    cls.writers[writer] = [ExcelWriter.getGroupShet('default_book', sheet)]
                cls.newid += 1
                try:
                    old_title = sheet.title
                    sheet.title = str(cls.newid) + '.' + cleanedSheetName(writer)
                except:
                    sheet.title = old_title
            else:
                xls = ExcelWriter(group='writer')
                cls.writers[writer] = [ExcelWriter.getGroupShet('writer', xls.sheets[0])]
                cls.newid += 1
                try:
                    xls.setFileName(str(cls.newid) + '.'  + cleanedFileName(writer) + '.xlsx')
                except:
                    xls.file = str(cls.newid) + '.xlsx'
            cls.writers['lock'].release()
            if new_model != None:
                if 'title_row' in new_model:
                    shet = cls.writers[writer][0]
                    shet['lock'].acquire()
                    for i in range(new_model['title_row']+1):
                        shet['xls'].updateRowValues(shet['sheet'], i, new_model['xls'].getRowValues(new_model['sheet'], i, check_merge=False))
                    exception = ['xls', 'sheet', 'lock']
                    exception.extend(not_copy)
                    for name in new_model:
                        if not (name in exception):
                            shet[name] = new_model[name]
                    for name in auto_set:
                        shet[name] = auto_set[name]
                    shet['lock'].release()
        else:
            cls.writers['lock'].release()
        return cls.writers[writer]
    @classmethod
    def loadWriter(cls, shet, writer):
        cls.writers['lock'].acquire()
        if writer in cls.writers:
            cls.writers[writer].append(shet)
        else:
            cls.writers[writer] = [shet]
        cls.writers['lock'].release()
    @classmethod
    def resetWriters(cls):
        pops = []
        cls.writers['lock'].acquire()
        cls.newid = 0
        for writer in cls.writers:
            if writer != 'lock':
                pops.append(writer)
        for writer in pops:
            cls.writers.pop(writer)
        cls.writers['lock'].release()
    @classmethod
    def createWriterFromGroup(cls, group='writer', skipmodel=False):
        cls.writers['lock'].acquire()
        if group in cls.groups:
            cls.resetWriters()
            for shet in cls.groups[group]:
                if shet['valid']:
                    if skipmodel:
                        if shet['sheet'] is cls.model['sheet']:
                            continue
                    writer = createStringFromVarList(shet['xls'].getRowValuesByPos(shet['sheet'], shet['title_row']+1, shet['split_cols'], check_merge=False), sep='-', add_quotes=False, skip_none=True)
                    if len(writer) == 0:
                        writer = shet['xls'].getSheetName(shet['sheet'])
                        if cls.shtre.match(writer) != None:
                            writer = shet['xls'].getBookName(True)
                    cls.loadWriter(shet, writer)
        else:
            return False
        cls.writers['lock'].release()
    @classmethod
    def resetBook(cls):
        cls.book = None
    @classmethod
    def saveBook(cls):
        if cls.book != None:
            cls.book.save()

if __name__ != "__main__":
    Cell.classInit()
    ExcelReader.classInit()
    ExcelWriter.classInit()
    